#!/usr/bin/env python
# coding: utf-8

# In[1]:


#Manual:
# Step 1) Create reddit-profile and register for developer app.
    # See how to on: https://www.youtube.com/watch?v=FdjVoOf9HN4 until minute 3.
    # Insert values for client_id, secret_key, and profile_name
#Step 2) Execute all libraries
#Step 3) Insert path to model (location of trained model) as well as tokenizer then execute
#Step 4) Execute all lists and functions
CLIENT_ID = 'zAiWbboKu9eCWztcBjxAwQ'
SECRET_KEY = 'f0DgQeim1r89SJEXxrY49r5Z7ypQ2Q'
PROFILE_NAME = 'Pink_Strawman'


# In[2]:


import praw
import xlsxwriter
import openpyxl
from pathlib import Path
import os
import time
from operator import itemgetter
from collections import Counter
from tqdm.notebook import tqdm
import pandas as pd
from torch.nn.functional import softmax
import torch.nn as nn
import torch
from transformers import DistilBertTokenizerFast, AutoTokenizer, AutoModelForSequenceClassification, DistilBertForSequenceClassification, AdamW, DistilBertConfig, get_linear_schedule_with_warmup
import sys, time, datetime, random
from keras.preprocessing.sequence import pad_sequences
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from torch.utils.data import TensorDataset, DataLoader, RandomSampler, SequentialSampler
import numpy as np
sys.path.append("/Users/raoultschudin/Desktop/Dokumente_Uni/MA/Model/reddit_incivility-master/scripts")
from CustomIterableDataset import CustomIterableDataset
from load_data import get_data


# In[3]:


#assign developer_profile data
reddit = praw.Reddit(client_id=CLIENT_ID, client_secret=SECRET_KEY, user_agent=PROFILE_NAME)
#assign model (insert path with model)
model = DistilBertForSequenceClassification.from_pretrained('/Users/raoultschudin/Desktop/Dokumente_Uni/MA/Model/reddit_incivility-master/scripts/Inciv_model')
#assign tokenizer
#config = BertConfig.from_json_file('../models/bert_classifier_2epoch_256size/config.json')
tokenizer = DistilBertTokenizerFast.from_pretrained('distilbert-base-uncased')


# In[7]:


#Create lists of all variables (needed for writing xls file with dictionary)
#---------------------------------------------------------------------#
##!!!ONLY REFRESH IF STARTING OR IF CURRENT DATA SHOULD BE DELETED!!!##
#---------------------------------------------------------------------#
list_topic = []; list_topic.clear()
list_title = []; list_title.clear()
depth_list_topic=[]; depth_list_topic.clear()
width_list_topic = []; width_list_topic.clear()
h_index_list_topic = []; h_index_list_topic.clear()
r_list_topic = []; r_list_topic.clear()
time_period = []; time_period.clear()
total_comments = []; total_comments.clear()
unique_users_list_topic = []; unique_users_list_topic.clear()
participation_rate = []; participation_rate.clear()
participation_rate_list = []; participation_rate_list.clear()
total_length_of_post_list = []; total_length_of_post_list.clear()
average_chars_list_topic = []; average_chars_list_topic.clear()
num_of_chars_list = []; num_of_chars_list.clear()
num_of_rules_list = []; num_of_rules_list.clear()
main_rules_list = []; main_rules_list.clear()
rules_description_list = []; rules_description_list.clear()
chars_per_rule_corpus_list = []; chars_per_rule_corpus_list.clear()
author_list = []; author_list.clear()
karma_of_author = []; karma_of_author.clear()
author_is_moderator = []; author_is_moderator.clear()
term_membership = []; term_membership.clear()
avg_term_membership = []; avg_term_membership.clear()
homogeneity = []; homogeneity.clear()
comment_list = []; comment_list.clear()
subreddit_list = []; subreddit_list.clear()
subreddit_id_list = []; subreddit_id_list.clear()
top_comment_list = []; top_comment_list.clear()
upvotes_list = []; upvotes_list.clear()
avg_upvotes = []; avg_upvotes.clear()
incivility_prob_list = []; incivility_prob_list.clear()
civility_prob_list = []; civility_prob_list.clear()
incivility_nom_list = []; incivility_nom_list.clear()
civility_nom_list = []; civility_nom_list.clear()
empty = []


# In[5]:


#execute

def process_comment(comment, depth=0):
    #Generate comment bodies and depths.
    yield comment.body, depth
    for reply in comment.replies:
        yield from process_comment(reply, depth + 1)

def get_post_comments(post, more_limit=None):
    #Get a list of (body, depth) pairs for the comments in the post.
    comments = []
    post.comments.replace_more(limit=more_limit)
    for top_level in post.comments:
        comments.extend(process_comment(top_level))
        time.sleep(0.01)
    return comments

def check_post2(post):
    #recursive function to determine the unique_users per post
    for comment in post:
        if comment.author not in unique_users2:
            unique_users2.append(comment.author)
            check_post2(comment.replies)

def check_post3(post):
    for comment in post:
        rating.append(comment.score)
        check_post3(comment.replies)

def check_post4(post):
    for comment in post:
        global none_counter
        try:
            if comment.author == None:
                terms_of_membership.append(int(0))
                none_counter += 1
            else:
                terms_of_membership_sec = (time.time() - comment.author.created_utc)
                terms_of_membership_min = terms_of_membership_sec/60
                terms_of_membership_hour = terms_of_membership_min/60
                terms_of_membership_day = terms_of_membership_hour/24
                terms_of_membership_year = terms_of_membership_day/365
                terms_of_membership.append(terms_of_membership_year)
        except AttributeError:
            terms_of_membership.append(int(0))
            none_counter += 1
            pass
        check_post4(comment.replies)

def metrics(df):
    truepos = []
    falsepos = []
    falseneg = []
    df = df.reset_index(drop=True)
    for n in range(len(df)):
        if df["correction"][n] == 1:
            if df["classification"][n] == 1:
                truepos.extend(df["comment"][n])
            if df["classification"][n] != 1:
                falseneg.extend(df["comment"][n])
        if df["correction"][n] == 0:
            if df["classification"][n] == 1:
                falsepos.extend(df["comment"][n])
            if df["classification"][n] == 0:
                truepos.extend(df["comment"][n])
    precision = len(truepos) / (len(truepos) + len(falsepos))
    recall = len(truepos) / (len(truepos) + len(falseneg))
    print("comments tested: ", len(df))
    print(f"---\nAccuracy: {accuracy_score(df['correction'], df['classification'])}")
    print(f"Precision: {round(precision,3)}")
    print(f"Recall: {round(recall,3)}")
    
    
def analyze_all_comments(subm, bar, post, ninetieth_percentile):
    for comments in subm:
        
        #instruction for how to print comments when executed
        global comment
        comment += 1
        if isinstance(post, int):
            print("Comment", comment, "in post", post, "with topic", "*", search_term,"*:\n", comments.body, "\n")
        else:
            print("Comment", comment, "in post", post, ":\n", comments.body, "\n")
        
        #comment_body
        comment_list.append(comments.body)
        
        #author, author_karma & author_is_moderator
        try:
            if comments.author == None:
                author_list.append('Deleted_account')
                author_is_moderator.append(int(0))
                karma_of_author.append(0)
                term_membership.append('None')
                avg_term_membership.append(avg_term_of_membership)
            else:
                term_membership_sec = (time.time() - comments.author.created_utc)
                term_membership_min = term_membership_sec/60
                term_membership_hour = term_membership_min/60
                term_membership_day = term_membership_hour/24
                term_membership_year = term_membership_day/365
                author_list.append(comments.author.name)
                karma_of_author.append(comments.author.comment_karma)
                term_membership.append(term_membership_year)
                avg_term_membership.append(avg_term_of_membership)
                if comments.author.is_mod:
                    print("written by a moderator")
                    author_is_moderator.append(int(1))
                else:
                    author_is_moderator.append(int(0))
        except AttributeError:
            author_list.append('Suspended_account')
            author_is_moderator.append(int(0))
            karma_of_author.append(0)
            term_membership.append('None')
            avg_term_membership.append(avg_term_of_membership)
            pass
    
        #subreddit
        subreddit_list.append(str(comments.subreddit))
        
        #subreddit_id
        subreddit_id_list.append(comments.subreddit.name)
        
        #time-difference between creation of subreddit and data-collection
        second_diff = (time.time() - comments.subreddit.created_utc)
        minute_diff = (second_diff/60)
        hour_diff = (minute_diff/60)
        day_diff = (hour_diff/24)
        year_diff = day_diff/365
        time_period.append(year_diff)
        
        #amount_of_rules, main_rules, rule_description
        rule_counter = 0
        single_rules_list1 = []
        single_rules_list2 = []
        for rule in comments.subreddit.rules:
            rule_counter += 1
            single_rules_list1.append(str(rule.description))
            single_rules_list2.append(str(rule))
        single_rules_string1 = '\n'.join(single_rules_list1)
        single_rules_string2 = '\n'.join(single_rules_list2)
        chars_per_rule_corpus_list.append(len(single_rules_string1))
        rules_description_list.append(single_rules_string1)
        main_rules_list.append(single_rules_string2)
        num_of_rules_list.append(rule_counter)
        
        #num_of_characters
        num_of_chars_list.append(len(comments.body))
        
        #total_comments
        total_comments.append(commentcount2)
        
        #participation_rate
        participation_ratio = len(unique_users2)/commentcount2
        participation_rate.append(participation_ratio)
        
        #homogeneity
        homogeneity_score = ((avg_score * avg_term_of_membership)/participation_ratio)
        homogeneity.append(homogeneity_score)
        
        #topic
        replaced_search_term = search_term.replace('"',"")
        list_topic.append(replaced_search_term)
        
        #title
        list_title.append(submission.title)
        
        #depth
        depth_list_topic.append(depth)
        
        #width & level of width
        width_list_topic.append(width2[0])
        
        #h-index
        h_index_list_topic.append(h_index2)
        
        #r-value
        r_list_topic.append(r)
        
        #upvotes
        upvotes_list.append(comments.score)
        
        #average_upvotes_in_post
        avg_upvotes.append(avg_score)
        
        #top-comment
        if comments.score > ninetieth_percentile:
            top_comment_list.append(int(1))
        else:
            top_comment_list.append(int(0))
        
        #total length of submission
        total_length_of_post_list.append(sumofchars2)
        
        #number_of_unique_users
        unique_users_list_topic.append(len(unique_users2))
        
        #average_characters_per_comment
        average_chars_list_topic.append(avg_chars_per_comment2)
        
        #incivility
            #get scores for the inserted strings (here 'comments.body')
                #if string is too long (>512 tokens)
                    #string is splitted into multiple chunks and their predicted value is averaged such that you receive one probability also for longer strings
            #at the end two variables are generated:
                #(in-)civility_prob: Probability of (in-)civility of the comments
                #(in-)civility_nom: Assignment of (in-)civility to either 1 (if p ≥ 0.5) or 0 (if p < 0.5)
        inputs = tokenizer.encode_plus(comments.body, add_special_tokens=False, return_tensors="pt")
        input_id_chunks = inputs['input_ids'][0].split(510)
        mask_chunks = inputs['attention_mask'][0].split(510)
        for tensor in input_id_chunks:
            a = torch.arange(10)
        torch.cat([torch.Tensor([101]), a, torch.Tensor([102])])
        chunksize = 512
        input_id_chunks = list(input_id_chunks)
        mask_chunks = list(mask_chunks)
        for i in range(len(input_id_chunks)):
            input_id_chunks[i] = torch.cat([torch.Tensor([101]), input_id_chunks[i], torch.Tensor([102])])
            mask_chunks[i] = torch.cat([torch.Tensor([1]), mask_chunks[i], torch.Tensor([1])])
            pad_len = chunksize - input_id_chunks[i].shape[0]
            if pad_len > 0:
                input_id_chunks[i] = torch.cat([input_id_chunks[i], torch.Tensor([0] * pad_len)])
                mask_chunks[i] = torch.cat([mask_chunks[i], torch.Tensor([0] * pad_len)])
        input_ids = torch.stack(input_id_chunks)
        attention_mask = torch.stack(mask_chunks)
        input_dict = {
                'input_ids': input_ids.long(),
                'attention_mask': attention_mask.int()
                }
        input_dict
        outputs = model(**input_dict)
        outputs = outputs[0]
        hate_probabilities = torch.nn.functional.softmax(outputs, dim=1)
        mean_hate_probabilities = hate_probabilities.mean(dim=0)
        mean_hate_probabilities = hate_probabilities.tolist()[0]
        hate_speech_prob = mean_hate_probabilities[1]
        normal_prob = mean_hate_probabilities[0]
        incivility_prob_list.append(hate_speech_prob)
        civility_prob_list.append(normal_prob)
        if hate_speech_prob >= 0.5:
            incivility_nom_list.append(int(1))
            civility_nom_list.append(int(0))
        else:
            incivility_nom_list.append(int(0))
            civility_nom_list.append(int(1))
        
        #Update progress bar and initialize recursion (to get values for comments of comments of comments of...)
        bar.update(1)
        time.sleep(0.01)
        analyze_all_comments(comments.replies, bar, post, ninetieth_percentile)


# In[6]:


# execute to analyze comments in a single or several subreddits matching the search-tearm
#if search term contains more than one word use "" to indicate the whole search-term.
# e.g.:
    #- one-word search: 'Climate'
    #- multiple-word search: '"Climate Change"', '"The Climate Change"', etc.

search_term = 'Trump OR Biden'

#subredit_name:
    #subreddits can be looked for specifically and can also be added together:
        # e.g.:
            #- Single subreddit: "news";
            #- Multiple subreddits: "news+music+sports" (where "news", "music", and "sports" are independent subreddits);
            #- All subreddits: "all"

subreddit_name = 'PoliticalDiscussion'

#restrictions:
    #sort_restriction:
        #assign under which order the comments will be analyzed
        #e.g.:
            #'relevance', 'hot', 'top', 'new', 'comments'
    #time_restriction:
        #determine time-horizon of eligible comments
        #e.g.:
            #'all', 'year', 'month', 'week', 'day', 'hour'
        #attention: Time to process all comments can increase extremely if chosing a too wide time-horizon  

sort_restriction = 'comments'
time_restriction = 'year'

post = 0
counter = 0

for submission in reddit.subreddit(subreddit_name).search(search_term, sort=sort_restriction, time_filter=time_restriction, limit=None):
    counter += 1
for submission in tqdm(reddit.subreddit(subreddit_name).search(search_term, sort=sort_restriction, time_filter=time_restriction, limit=None), total=counter):
    if not get_post_comments(submission)==empty and len(get_post_comments(submission))<=400 and len(get_post_comments(submission))>=100:
        post +=1
        level_counter = 0
        global none_counter
        none_counter = 0
        unique_users2 = []
        rating = []
        terms_of_membership = []
        sumofchars2=0
        commentcount2=0
        print("Counting comments...")
        for comments in get_post_comments(submission):
            sumofchars2+=len(comments[0])
            commentcount2+=1
        avg_chars_per_comment2=(sumofchars2/commentcount2)
        print("Computing depth...")
        depth = max(get_post_comments(submission), key=itemgetter(1))[1]+1
        print("Computing width...")
        Width_list2=[i[1] for i in get_post_comments(submission)]
        Max_occurrence_lst2 = Counter(Width_list2).most_common(1)
        level_of_max_comments2 = [i[0] for i in Max_occurrence_lst2]
        width2=[i[1] for i in Max_occurrence_lst2]
        print("Computing h-index and r-value...")
        CommentsPerLevel_List2 = Counter(Width_list2).most_common()
        h_index2 = 0
        sorted_by_level2 = sorted(CommentsPerLevel_List2, key=lambda tup: tup[0])
        for level in sorted_by_level2:
            if level[0] < level[1]:
                h_index2 +=1
        r = h_index2+(1/submission.num_comments)
        print("Determining number of unique users...")
        check_post2(submission.comments)
        print("Determining scores...")
        check_post3(submission.comments)
        sum_score = sum(rating)
        avg_score = (sum_score/(commentcount2))
        print("Determining average membership duration...")
        check_post4(submission.comments)
        sum_terms_of_membership = sum(terms_of_membership)
        avg_term_of_membership = (sum_terms_of_membership/(commentcount2-none_counter))
        print("Determining top comments...")
        prepare_rating = np.array(rating)
        ninetieth_percentile = np.percentile(prepare_rating, 90)
        global comment
        comment = 0
        bar = tqdm(total=commentcount2)
        print("!Starting comment evaluation!")
        analyze_all_comments(submission.comments, bar, post, ninetieth_percentile)
        print("Comment evaluation over")


# In[99]:


#execute to analyze comments in specific post with link
search_term = 'None'
link = "https://www.reddit.com/r/birding/comments/tipqc8/what_is_this_bird_seen_in_zurich_switzerland/"
submission = reddit.submission(url=link)
if not get_post_comments(submission)==empty:
    global none_counter
    none_counter = 0
    post = link
    unique_users2 = []
    rating = []
    terms_of_membership = []
    sumofchars2=0
    commentcount2=0
    for comments in get_post_comments(submission):
        sumofchars2+=len(comments[0])
        commentcount2+=1
    avg_chars_per_comment2=(sumofchars2/commentcount2)
    depth = max(get_post_comments(submission), key=itemgetter(1))[1]+1
    Width_list2=[i[1] for i in get_post_comments(submission)]
    Max_occurrence_lst2 = Counter(Width_list2).most_common(1)
    level_of_max_comments2 = [i[0] for i in Max_occurrence_lst2]
    width2=[i[1] for i in Max_occurrence_lst2]
    CommentsPerLevel_List2 = Counter(Width_list2).most_common()
    h_index2 = 0
    sorted_by_level2 = sorted(CommentsPerLevel_List2, key=lambda tup: tup[0])
    for level in sorted_by_level2:
        time.sleep(0.01)
        if level[0] < level[1]:
            h_index2 +=1
    r = h_index2+(1/submission.num_comments)
    check_post2(submission.comments)
    check_post3(submission.comments)
    check_post4(submission.comments)
    sum_score = sum(rating)
    avg_score = (sum_score/commentcount2)
    prepare_rating = np.array(rating)
    sum_terms_of_membership = sum(terms_of_membership)
    print(none_counter)
    avg_term_of_membership = (sum_terms_of_membership/(commentcount2-none_counter))
    ninetieth_percentile = np.percentile(prepare_rating, 90)
    global comment
    comment = 0
    bar = tqdm(total=commentcount2)
    analyze_all_comments(submission.comments, bar, post, ninetieth_percentile)


# In[7]:


#create dictionary with variable_lists to prepare the data for the creation of the xlsx-file
variable_dict_topic = {
                'topic': list_topic,
                'subreddit': subreddit_list,
                'subreddit_id': subreddit_id_list,
                'title': list_title,
                'time_period': time_period,
                'total_comments': total_comments,
                'unique_users': unique_users_list_topic,
                'participation_rate': participation_rate,
                'total_post_length': total_length_of_post_list,
                'avg_num_of_chars': average_chars_list_topic,
                'avg_score': avg_upvotes,
                'avg_term_membership': avg_term_membership,
                'homogeneity_score': homogeneity,
                'depth': depth_list_topic,
                'width': width_list_topic,
                'h_index': h_index_list_topic,
                'r-value': r_list_topic,
                'main_rules': main_rules_list,
                'num_of_rules': num_of_rules_list,
                'description_rules': rules_description_list,
                'chars_per_rule_corpus': chars_per_rule_corpus_list,
                'author': author_list,
                'body': comment_list,
                'score': upvotes_list,
                'top_comment': top_comment_list,
                'num_of_chars': num_of_chars_list,
                'karma': karma_of_author,
                'term_membership': term_membership,
                'moderator': author_is_moderator,
                'incivility_prob_':incivility_prob_list,
                'civility_prob_list': civility_prob_list,
                'incivility_nom': incivility_nom_list,
                'civility_nom': civility_nom_list}


# In[8]:


#create excel-file and insert values with created list above
#assign path where to create xlsx-file
#if existing file should not be overwriten with current dictionary, change the name of the file or path
workbook = xlsxwriter.Workbook('Desktop/Dokumente_Uni/MA/PythonOutput/variables_deliberation_PoliticalDiscussion_TrumpORBiden.xlsx', options={'strings_to_urls': False})
worksheet = workbook.add_worksheet()
col_num = 0
for key, value in variable_dict_topic.items():
    worksheet.write(0, col_num, key)
    worksheet.write_column(1, col_num, value)
    col_num += 1
workbook.close()


# In[62]:


#get test-scores for comments
#insert path with test-file, has to be xlsx format

xlsx_file = Path('/Users/raoultschudin/Desktop/Dokumente_Uni/MA/PythonOutput/variables_deliberation_test1.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active
auto_values = []
manual_values = []
tested_comments = []

#insert column in xlsx-file containing the comments to be tested, the estimated values, and the self-determined values
for row in sheet.iter_rows():
    tested_comments.append(row[4].value)
    auto_values.append(row[14].value)
    manual_values.append(row[15].value)

#create dataframe object for test-score calculation
#if first row in excel-file starts with variable-names, leave code like this, otherwise take out slicing
dataframe_dict = {
                'comment': tested_comments[1:],
                'correction': auto_values[1:],
                'classification': manual_values[1:]}

df = pd.DataFrame(data=dataframe_dict)

metrics(df)


# In[ ]:




